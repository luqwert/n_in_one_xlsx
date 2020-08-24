#!/usr/bin/env python
# _*_ coding:utf-8 _*_
# @Author  : lusheng
import os
from openpyxl import load_workbook


def get_dir():  # 获取当前文件夹下非总表的excel文件列表
    dir_list = []
    for root, dirs, files in os.walk('./'):
        for file in files:
            if os.path.splitext(file)[1] == '.xlsx' and file != '总表.xlsx':
                dir_list.append(file)
    return dir_list


def get_sheets(dir):
    wb_fenbiao = load_workbook(dir, data_only=False)
    sheet_list = wb_fenbiao.sheetnames
    print(sheet_list)
    return wb_fenbiao, sheet_list


def get_rows_data(wb_fenbiao, sheet):
    rowsdata = []
    max_row = wb_fenbiao[sheet].max_row
    max_column = wb_fenbiao[sheet].max_column
    if wb_fenbiao[sheet].cell(row=3, column=5).value is None:
        return rowsdata
    elif '合同额' in wb_fenbiao[sheet].cell(row=3, column=5).value:
        for m in range(5, max_row + 1):
            cells = []

            if wb_fenbiao[sheet].cell(row=m, column=3).value == None:
                pass
            else:
                # print(wb_fenbiao[sheet].cell(row=m, column=1).value)
                for n in range(1, max_column):
                    cells.append(wb_fenbiao[sheet].cell(row=m, column=n).value)  # 获取data单元格数据
                rowsdata.append(cells)
    return rowsdata


def main():
    # 总表名称
    excel_path = '总表.xlsx'
    # 打开已经存在的表格并实例化，准备进行修改操作
    wb = load_workbook(excel_path)
    zongbiao = wb["Sheet1"]
    # n_of_rows = zongbiao.max_row
    n_of_cols = zongbiao.max_column
    n_of_rows = 5
    # print(n_of_rows,n_of_cols)
    dir_list = get_dir()
    print(dir_list)
    for dir in dir_list:
        wb_fenbiao, sheets = get_sheets(dir)
        for sheet in sheets:
            rowsdata = get_rows_data(wb_fenbiao, sheet)
            for row in rowsdata:
                for n in range(0, len(row)):
                    zongbiao.cell(row=n_of_rows, column=n + 1, value=row[n])
                n_of_rows = n_of_rows + 1
            wb.save('总表.xlsx')

    wb.close()


if __name__ == '__main__':
    main()
